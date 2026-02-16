"""Tracker data processor for student progress tracking.

Transforms raw CSV form responses into a comprehensive Excel workbook
with multiple tabs for different priority levels and a summary dashboard.
"""

import csv
import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.file_processor import FileProcessor, ProcessingResult


# ==================== Data Classes ====================

@dataclass
class StudentRecord:
    """Represents a processed student record with all calculated fields."""
    # Core identifiers
    student_id: str = ""
    name: str = ""
    member_id: str = ""
    discord_username: str = ""
    
    # Time tracking
    week: int = 0
    submission_date: str = ""
    wed_submitted: bool = False
    sun_submitted: bool = False
    submission_count_cumulative: int = 0
    
    # Phase tracking
    current_phase: str = ""
    weeks_in_phase: int = 1
    contribution_num: int = 1
    contribution_start_week: int = 1
    weeks_on_contribution: int = 1
    weeks_remaining: int = 8
    timeline_type: str = "Standard"
    phase_changed_this_week: bool = False
    
    # Links
    readme_link: str = ""
    issue_url: str = ""
    fork_url: str = ""
    mr_url: str = ""
    
    # Deliverables completion
    why_chosen_complete: bool = False
    reproduction_complete: bool = False
    solution_complete: bool = False
    implementation_complete: bool = False
    testing_complete: bool = False
    feedback_complete: bool = False
    deliverables_expected: int = 0
    deliverables_complete: int = 0
    
    # Git activity
    commits_this_week: int = 0
    last_commit_date: str = ""
    days_since_commit: int = 0
    total_commits: int = 0
    
    # MR tracking
    mr_status: str = ""
    mr_created_date: str = ""
    comment_count: int = 0
    has_maintainer_feedback: bool = False
    
    # Progress
    progress_summary: str = ""
    next_week_plan: str = ""
    blocked: bool = False
    blocker_desc: str = ""
    support_requested: str = ""
    
    # Issue tracking
    issue_url_previous_week: str = ""
    issue_changed: bool = False
    issue_change_week: int = 0
    issue_swap_detected: bool = False
    new_contribution_detected: bool = False
    
    # Grading and intervention
    grade_status: str = "🟢 ON TRACK"
    intervention_type: str = ""
    intervention_sent_date: str = ""
    consecutive_misses: int = 0
    
    # Office hours
    tue_office_hours: bool = False
    thu_office_hours: bool = False
    wed_lecture: bool = False
    
    # Notes
    cam_notes: str = ""
    
    # Raw data for reference
    raw_data: Dict[str, Any] = field(default_factory=dict)


# ==================== Column Mappings ====================

# Maps CSV column headers to StudentRecord fields
CSV_COLUMN_MAP = {
    "#": "student_id",
    "What's your name?": "name",
    "What's your Member ID?": "member_id",
    "Which week is this?": "week",
    "Which contribution are you reporting on?": "contribution_num",
    "Link to your contribution README": "readme_link",
    "Which submission are you completing?": "_submission_type",
    "What phase are you currently in?": "current_phase",
    "Direct link to your GitLab issue": "issue_url",
    "Have you completed the \"Why I chose this issue\" section in your README?": "why_chosen_complete",
    "Direct link to your GitLab fork": "fork_url",
    "Have you documented your reproduction process in your README?": "reproduction_complete",
    "Have you documented your solution approach in your README?": "solution_complete",
    "Have you documented your implementation progress in your README?": "implementation_complete",
    "Have you documented your testing strategy in your README?": "testing_complete",
    "Direct link to your Merge Request (MR) or Pull Request (PR)": "mr_url",
    "Have you documented any maintainer feedback in your README?": "feedback_complete",
    "Briefly summarize what you accomplished this week": "progress_summary",
    "What's your plan for next week?": "next_week_plan",
    "Are you currently blocked or stuck?": "blocked",
    "Describe what you're blocked on": "blocker_desc",
    "Submit Date (UTC)": "submission_date",
    "Tags": "_tags",
}


# ==================== Style Definitions ====================

class Styles:
    """Excel style definitions."""
    
    # Fills
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    RED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    ORANGE_FILL = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    LIGHT_YELLOW_FILL = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    DASHBOARD_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    DASHBOARD_SECTION_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
    
    # Fonts
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BOLD_FONT = Font(bold=True)
    TITLE_FONT = Font(bold=True, size=14)
    DASHBOARD_TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
    
    # Alignment
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Border
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


# ==================== Tracker Processor ====================

class TrackerDataProcessor(FileProcessor):
    """Processes tracker CSV data into a comprehensive Excel workbook.
    
    Creates multiple tabs:
    - Tab 1: Master Tracker (all fields)
    - Tab 2: At Risk Students (P1 Priority)
    - Tab 3: Flagged Students (P2 Priority)  
    - Tab 4: On Track Students (P3 Spot Checks)
    - Tab 5: Weekly Summary Dashboard
    """
    
    # Output column definitions for each tab
    ALL_COLUMNS = [
        "student_id", "name", "member_id", "discord_username", "week",
        "submission_date", "wed_submitted", "sun_submitted", "submission_count_cumulative",
        "current_phase", "weeks_in_phase", "contribution_num", "contribution_start_week",
        "weeks_on_contribution", "weeks_remaining", "timeline_type", "phase_changed_this_week",
        "readme_link", "issue_url", "fork_url", "mr_url",
        "why_chosen_complete", "reproduction_complete", "solution_complete",
        "implementation_complete", "testing_complete", "feedback_complete",
        "deliverables_expected", "deliverables_complete",
        "commits_this_week", "last_commit_date", "days_since_commit", "total_commits",
        "mr_status", "mr_created_date", "comment_count", "has_maintainer_feedback",
        "progress_summary", "next_week_plan", "blocked", "blocker_desc", "support_requested",
        "issue_url_previous_week", "issue_changed", "issue_change_week",
        "issue_swap_detected", "new_contribution_detected",
        "grade_status", "intervention_type", "intervention_sent_date", "consecutive_misses",
        "tue_office_hours", "thu_office_hours", "wed_lecture", "cam_notes"
    ]
    
    AT_RISK_COLUMNS = [
        "name", "week", "current_phase", "weeks_in_phase", "timeline_type",
        "sun_submitted", "consecutive_misses", "deliverables_complete",
        "commits_this_week", "blocked", "intervention_type", "readme_link", "cam_notes"
    ]
    
    FLAGGED_COLUMNS = [
        "name", "week", "current_phase", "weeks_in_phase", "timeline_type",
        "deliverables_complete", "commits_this_week", "days_since_commit",
        "blocked", "intervention_type", "readme_link"
    ]
    
    ON_TRACK_COLUMNS = [
        "name", "week", "current_phase", "weeks_in_phase", "submission_count_cumulative",
        "mr_status", "progress_summary", "cam_notes"
    ]
    
    @property
    def input_type(self) -> str:
        return "csv"
    
    @property
    def output_type(self) -> str:
        return "xlsx"
    
    def process(self, data: bytes, options: Optional[Dict[str, Any]] = None) -> ProcessingResult:
        """Process CSV data into multi-tab Excel workbook."""
        options = options or {}
        
        try:
            # Parse CSV
            text_data = data.decode('utf-8-sig')
            csv_reader = csv.DictReader(io.StringIO(text_data))
            raw_rows = list(csv_reader)
            
            if not raw_rows:
                return ProcessingResult(
                    success=False,
                    error_message="CSV file is empty"
                )
            
            # Transform to StudentRecord objects
            students = self._transform_records(raw_rows)
            
            # Calculate derived fields
            self._calculate_derived_fields(students)
            
            # Determine grade status and interventions
            self._calculate_grade_status(students)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create tabs (Master first, then priority tabs)
            self._create_master_tab(wb, students)
            self._create_at_risk_tab(wb, students)
            self._create_flagged_tab(wb, students)
            self._create_on_track_tab(wb, students)
            self._create_summary_tab(wb, students)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return ProcessingResult(
                success=True,
                output_data=output.read(),
                output_filename="tracker_report.xlsx",
                rows_processed=len(students)
            )
            
        except Exception as e:
            return ProcessingResult(
                success=False,
                error_message=f"Processing error: {e}"
            )
    
    def _transform_records(self, raw_rows: List[Dict]) -> List[StudentRecord]:
        """Transform raw CSV rows into StudentRecord objects."""
        students = []
        
        for row in raw_rows:
            student = StudentRecord()
            student.raw_data = row
            
            for csv_col, field_name in CSV_COLUMN_MAP.items():
                if csv_col in row:
                    value = row[csv_col]
                    
                    # Handle special field mappings
                    if field_name == "_submission_type":
                        if "Wednesday" in value:
                            student.wed_submitted = True
                        elif "Sunday" in value:
                            student.sun_submitted = True
                    elif field_name == "_tags":
                        # Check for AI Generated tag
                        if "AI Generated" in str(value):
                            student.cam_notes = "[AI Generated Response]"
                    elif field_name == "week":
                        # Extract week number
                        try:
                            week_str = value.replace("Week ", "").strip()
                            student.week = int(week_str)
                        except:
                            student.week = 0
                    elif field_name == "contribution_num":
                        # Extract contribution number
                        try:
                            if "Contribution" in value:
                                num = value.split()[-1]
                                student.contribution_num = int(num)
                            elif "No Contribution" in value:
                                student.contribution_num = 0
                        except:
                            student.contribution_num = 1
                    elif field_name == "current_phase":
                        # Normalize phase names
                        student.current_phase = self._normalize_phase(value)
                    elif field_name in ["why_chosen_complete", "reproduction_complete", 
                                       "solution_complete", "implementation_complete",
                                       "testing_complete", "feedback_complete"]:
                        # Convert 1/0 to boolean
                        setattr(student, field_name, str(value) == "1")
                    elif field_name == "blocked":
                        setattr(student, field_name, str(value) == "1")
                    elif hasattr(student, field_name):
                        setattr(student, field_name, value)
            
            students.append(student)
        
        return students
    
    def _normalize_phase(self, phase_str: str) -> str:
        """Normalize phase string to consistent format."""
        phase_str = str(phase_str).lower()
        
        if "1" in phase_str or "selection" in phase_str:
            return "Phase 1: Issue Selection"
        elif "2" in phase_str or "reproduction" in phase_str:
            return "Phase 2: Reproduction & Planning"
        elif "3" in phase_str or "implementation" in phase_str:
            return "Phase 3: Implementation"
        elif "4" in phase_str or "submission" in phase_str:
            return "Phase 4: Submission & Iteration"
        return phase_str
    
    def _calculate_derived_fields(self, students: List[StudentRecord]) -> None:
        """Calculate derived fields for each student."""
        for student in students:
            # Calculate deliverables expected based on phase
            phase_num = self._get_phase_number(student.current_phase)
            
            if phase_num >= 1:
                student.deliverables_expected = 1  # Why chosen
            if phase_num >= 2:
                student.deliverables_expected += 2  # Reproduction + Solution
            if phase_num >= 3:
                student.deliverables_expected += 2  # Implementation + Testing
            if phase_num >= 4:
                student.deliverables_expected += 1  # Feedback
            
            # Calculate deliverables complete
            student.deliverables_complete = sum([
                student.why_chosen_complete,
                student.reproduction_complete,
                student.solution_complete,
                student.implementation_complete,
                student.testing_complete,
                student.feedback_complete
            ])
            
            # Calculate weeks remaining (assuming 8-week program)
            student.weeks_remaining = max(0, 8 - student.week)
            
            # Determine timeline type
            if student.weeks_remaining < 3 and phase_num < 3:
                student.timeline_type = "Compressed"
            elif student.weeks_remaining < 2 and phase_num < 4:
                student.timeline_type = "Critical"
            else:
                student.timeline_type = "Standard"
    
    def _get_phase_number(self, phase_str: str) -> int:
        """Extract phase number from phase string."""
        if "1" in phase_str:
            return 1
        elif "2" in phase_str:
            return 2
        elif "3" in phase_str:
            return 3
        elif "4" in phase_str:
            return 4
        return 0
    
    def _calculate_grade_status(self, students: List[StudentRecord]) -> None:
        """Calculate grade status and intervention type for each student."""
        for student in students:
            phase_num = self._get_phase_number(student.current_phase)
            
            # Check for AT RISK conditions
            at_risk = False
            intervention = ""
            
            # Missing both submissions
            if not student.wed_submitted and not student.sun_submitted:
                at_risk = True
                intervention = "MISSING_BOTH"
            
            # Phase critical (stuck in early phase late in program)
            elif student.week >= 6 and phase_num <= 2:
                at_risk = True
                intervention = "PHASE_CRITICAL"
            
            # Stalled with blockers
            elif student.blocked and student.blocker_desc:
                at_risk = True
                intervention = "STALLED"
            
            # Check for FLAGGED conditions
            flagged = False
            
            # Missing deliverables for current phase
            if not at_risk:
                if student.deliverables_complete < student.deliverables_expected:
                    flagged = True
                    intervention = "MISSING_DELIVERABLES"
                
                # No recent commits
                elif student.days_since_commit > 7:
                    flagged = True
                    intervention = "NO_ACTIVITY"
                
                # Compressed timeline
                elif student.timeline_type == "Compressed":
                    flagged = True
                    intervention = "TIMELINE_COMPRESSED"
            
            # Set status
            if at_risk:
                student.grade_status = "🔴 AT RISK"
            elif flagged:
                student.grade_status = "🟡 FLAGGED"
            else:
                student.grade_status = "🟢 ON TRACK"
            
            student.intervention_type = intervention
    
    def _create_master_tab(self, wb: Workbook, students: List[StudentRecord]) -> None:
        """Create Tab 1: Master Sheet with all student data."""
        ws = wb.create_sheet("Master Tracker")
        
        # Define all columns in order
        headers = [
            "student_id", "name", "member_id", "discord_username", "week",
            "submission_date", "wed_submitted", "sun_submitted", "submission_count_cumulative",
            "current_phase", "weeks_in_phase", "contribution_num", "contribution_start_week",
            "weeks_on_contribution", "weeks_remaining", "timeline_type", "phase_changed_this_week",
            "readme_link", "issue_url", "fork_url", "mr_url",
            "why_chosen_complete", "reproduction_complete", "solution_complete",
            "implementation_complete", "testing_complete", "feedback_complete",
            "deliverables_expected", "deliverables_complete",
            "commits_this_week", "last_commit_date", "days_since_commit", "total_commits",
            "mr_status", "mr_created_date", "comment_count", "has_maintainer_feedback",
            "progress_summary", "next_week_plan", "blocked", "blocker_desc", "support_requested",
            "issue_url_previous_week", "issue_changed", "issue_change_week",
            "issue_swap_detected", "new_contribution_detected",
            "grade_status", "intervention_type", "intervention_sent_date", "consecutive_misses",
            "tue_office_hours", "thu_office_hours", "wed_lecture", "cam_notes"
        ]
        
        # Write header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = Styles.HEADER_FILL
            cell.font = Styles.HEADER_FONT
            cell.alignment = Styles.CENTER_ALIGN
            cell.border = Styles.THIN_BORDER
        
        # Write data rows
        for row_idx, student in enumerate(students, 2):
            data = [
                student.student_id,
                student.name,
                student.member_id,
                student.discord_username,
                student.week,
                student.submission_date,
                "✅" if student.wed_submitted else "❌",
                "✅" if student.sun_submitted else "❌",
                student.submission_count_cumulative,
                student.current_phase,
                student.weeks_in_phase,
                student.contribution_num,
                student.contribution_start_week,
                student.weeks_on_contribution,
                student.weeks_remaining,
                student.timeline_type,
                "✅" if student.phase_changed_this_week else "",
                student.readme_link,
                student.issue_url,
                student.fork_url,
                student.mr_url,
                "✅" if student.why_chosen_complete else "❌",
                "✅" if student.reproduction_complete else "❌",
                "✅" if student.solution_complete else "❌",
                "✅" if student.implementation_complete else "❌",
                "✅" if student.testing_complete else "❌",
                "✅" if student.feedback_complete else "❌",
                student.deliverables_expected,
                student.deliverables_complete,
                student.commits_this_week,
                student.last_commit_date,
                student.days_since_commit,
                student.total_commits,
                student.mr_status,
                student.mr_created_date,
                student.comment_count,
                "✅" if student.has_maintainer_feedback else "",
                student.progress_summary,
                student.next_week_plan,
                "🚫" if student.blocked else "",
                student.blocker_desc,
                student.support_requested,
                student.issue_url_previous_week,
                "✅" if student.issue_changed else "",
                student.issue_change_week if student.issue_change_week else "",
                "⚠️" if student.issue_swap_detected else "",
                "🆕" if student.new_contribution_detected else "",
                student.grade_status,
                student.intervention_type,
                student.intervention_sent_date,
                student.consecutive_misses,
                "✅" if student.tue_office_hours else "",
                "✅" if student.thu_office_hours else "",
                "✅" if student.wed_lecture else "",
                student.cam_notes
            ]
            
            # Determine row color based on grade status
            if student.grade_status == "🔴 AT RISK":
                row_fill = Styles.RED_FILL
            elif student.grade_status == "🟡 FLAGGED":
                row_fill = Styles.LIGHT_YELLOW_FILL
            else:
                row_fill = Styles.LIGHT_GREEN_FILL
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.border = Styles.THIN_BORDER
                cell.alignment = Styles.LEFT_ALIGN
        
        self._auto_fit_columns(ws)
        ws.freeze_panes = 'A2'
    
    def _create_at_risk_tab(self, wb: Workbook, students: List[StudentRecord]) -> None:
        """Create Tab 2: At Risk Students."""
        ws = wb.create_sheet("P1 - At Risk")
        
        # Filter students
        at_risk = [s for s in students if s.grade_status == "🔴 AT RISK"]
        
        # Sort by intervention priority
        priority_order = {"MISSING_BOTH": 0, "PHASE_CRITICAL": 1, "STALLED": 2}
        at_risk.sort(key=lambda s: priority_order.get(s.intervention_type, 99))
        
        # Write header
        headers = ["Name", "Week", "Phase", "Weeks in Phase", "Timeline", 
                   "Sun Submitted", "Consecutive Misses", "Deliverables",
                   "Commits", "Blocked", "Intervention Type", "README Link", "Notes"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = Styles.HEADER_FILL
            cell.font = Styles.HEADER_FONT
            cell.alignment = Styles.CENTER_ALIGN
            cell.border = Styles.THIN_BORDER
        
        # Write data
        for row_idx, student in enumerate(at_risk, 2):
            data = [
                student.name,
                student.week,
                student.current_phase,
                student.weeks_in_phase,
                student.timeline_type,
                "✅" if student.sun_submitted else "❌",
                student.consecutive_misses,
                f"{student.deliverables_complete}/{student.deliverables_expected}",
                student.commits_this_week,
                "🚫 " + student.blocker_desc[:30] if student.blocked else "",
                student.intervention_type,
                student.readme_link,
                student.cam_notes
            ]
            
            # Determine row color
            if student.intervention_type == "MISSING_BOTH":
                row_fill = Styles.RED_FILL
            elif student.intervention_type == "PHASE_CRITICAL":
                row_fill = Styles.ORANGE_FILL
            elif student.timeline_type in ["Compressed", "Critical"]:
                row_fill = Styles.YELLOW_FILL
            else:
                row_fill = Styles.RED_FILL
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.border = Styles.THIN_BORDER
                cell.alignment = Styles.LEFT_ALIGN
        
        self._auto_fit_columns(ws)
        ws.freeze_panes = 'A2'
    
    def _create_flagged_tab(self, wb: Workbook, students: List[StudentRecord]) -> None:
        """Create Tab 3: Flagged Students."""
        ws = wb.create_sheet("P2 - Flagged")
        
        # Filter students
        flagged = [s for s in students if s.grade_status == "🟡 FLAGGED"]
        
        # Sort by weeks in phase (descending)
        flagged.sort(key=lambda s: s.weeks_in_phase, reverse=True)
        
        # Write header
        headers = ["Name", "Week", "Phase", "Weeks in Phase", "Timeline",
                   "Deliverables", "Commits This Week", "Days Since Commit",
                   "Blocked", "Intervention Type", "README Link"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = Styles.HEADER_FILL
            cell.font = Styles.HEADER_FONT
            cell.alignment = Styles.CENTER_ALIGN
            cell.border = Styles.THIN_BORDER
        
        # Write data
        for row_idx, student in enumerate(flagged, 2):
            data = [
                student.name,
                student.week,
                student.current_phase,
                student.weeks_in_phase,
                student.timeline_type,
                f"{student.deliverables_complete}/{student.deliverables_expected}",
                student.commits_this_week,
                student.days_since_commit,
                "🚫 Blocked" if student.blocked else "",
                student.intervention_type,
                student.readme_link
            ]
            
            # Determine row color
            row_fill = Styles.ORANGE_FILL if student.blocked else Styles.LIGHT_YELLOW_FILL
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.border = Styles.THIN_BORDER
                cell.alignment = Styles.LEFT_ALIGN
        
        self._auto_fit_columns(ws)
        ws.freeze_panes = 'A2'
    
    def _create_on_track_tab(self, wb: Workbook, students: List[StudentRecord]) -> None:
        """Create Tab 4: On Track Students."""
        ws = wb.create_sheet("P3 - On Track")
        
        # Filter students
        on_track = [s for s in students if s.grade_status == "🟢 ON TRACK"]
        
        # Sort by week (descending)
        on_track.sort(key=lambda s: s.week, reverse=True)
        
        # Write header
        headers = ["Name", "Week", "Phase", "Weeks in Phase", "Submission Count",
                   "MR Status", "Progress Summary", "Notes"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = Styles.HEADER_FILL
            cell.font = Styles.HEADER_FONT
            cell.alignment = Styles.CENTER_ALIGN
            cell.border = Styles.THIN_BORDER
        
        # Write data
        for row_idx, student in enumerate(on_track, 2):
            # Add icons for achievements
            mr_display = student.mr_status
            if "merged" in student.mr_status.lower():
                mr_display = "⭐ " + mr_display
            
            notes = student.cam_notes
            if student.contribution_num >= 2:
                notes = "🏆 2nd Contribution! " + notes
            
            data = [
                student.name,
                student.week,
                student.current_phase,
                student.weeks_in_phase,
                student.submission_count_cumulative,
                mr_display,
                student.progress_summary[:100] + "..." if len(student.progress_summary) > 100 else student.progress_summary,
                notes
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = Styles.LIGHT_GREEN_FILL
                cell.border = Styles.THIN_BORDER
                cell.alignment = Styles.LEFT_ALIGN
        
        self._auto_fit_columns(ws)
        ws.freeze_panes = 'A2'
    
    def _create_summary_tab(self, wb: Workbook, students: List[StudentRecord]) -> None:
        """Create Tab 5: Weekly Summary Dashboard."""
        ws = wb.create_sheet("Weekly Summary")
        
        # Calculate statistics
        total = len(students)
        on_track = len([s for s in students if s.grade_status == "🟢 ON TRACK"])
        flagged = len([s for s in students if s.grade_status == "🟡 FLAGGED"])
        at_risk = len([s for s in students if s.grade_status == "🔴 AT RISK"])
        
        sun_submitted = len([s for s in students if s.sun_submitted])
        wed_submitted = len([s for s in students if s.wed_submitted])
        
        phase_dist = {1: 0, 2: 0, 3: 0, 4: 0}
        for s in students:
            phase_num = self._get_phase_number(s.current_phase)
            if phase_num in phase_dist:
                phase_dist[phase_num] += 1
        
        mr_submitted = len([s for s in students if s.mr_url])
        mr_merged = len([s for s in students if "merged" in s.mr_status.lower()])
        
        interventions_sent = len([s for s in students if s.intervention_type])
        
        # Get current week from data
        current_week = max(s.week for s in students) if students else 0
        
        # Create dashboard layout
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        
        # Title
        ws.merge_cells('B2:C2')
        title_cell = ws.cell(row=2, column=2, value=f"WEEK {current_week} OVERVIEW")
        title_cell.fill = Styles.DASHBOARD_HEADER_FILL
        title_cell.font = Styles.DASHBOARD_TITLE_FONT
        title_cell.alignment = Styles.CENTER_ALIGN
        
        # Total students
        row = 4
        ws.cell(row=row, column=2, value="Total Students:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=total)
        
        # Status breakdown
        row += 2
        ws.cell(row=row, column=2, value="🟢 On Track:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=f"{on_track} ({on_track/total*100:.1f}%)" if total else "0")
        ws.cell(row=row, column=2).fill = Styles.GREEN_FILL
        
        row += 1
        ws.cell(row=row, column=2, value="🟡 Flagged:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=f"{flagged} ({flagged/total*100:.1f}%)" if total else "0")
        ws.cell(row=row, column=2).fill = Styles.YELLOW_FILL
        
        row += 1
        ws.cell(row=row, column=2, value="🔴 At Risk:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=f"{at_risk} ({at_risk/total*100:.1f}%)" if total else "0")
        ws.cell(row=row, column=2).fill = Styles.RED_FILL
        
        # Submissions section
        row += 2
        ws.merge_cells(f'B{row}:C{row}')
        section = ws.cell(row=row, column=2, value="Submissions")
        section.fill = Styles.DASHBOARD_SECTION_FILL
        section.font = Styles.BOLD_FONT
        
        row += 1
        ws.cell(row=row, column=2, value="└─ Sunday:")
        ws.cell(row=row, column=3, value=f"{sun_submitted}/{total} ({sun_submitted/total*100:.1f}%)" if total else "0")
        
        row += 1
        ws.cell(row=row, column=2, value="└─ Wednesday:")
        ws.cell(row=row, column=3, value=f"{wed_submitted}/{total} ({wed_submitted/total*100:.1f}%)" if total else "0")
        
        # Phase distribution
        row += 2
        ws.merge_cells(f'B{row}:C{row}')
        section = ws.cell(row=row, column=2, value="Phase Distribution")
        section.fill = Styles.DASHBOARD_SECTION_FILL
        section.font = Styles.BOLD_FONT
        
        for phase in [1, 2, 3, 4]:
            row += 1
            ws.cell(row=row, column=2, value=f"└─ Phase {phase}:")
            ws.cell(row=row, column=3, value=f"{phase_dist[phase]} students")
        
        # MR section
        row += 2
        ws.cell(row=row, column=2, value="MRs Submitted:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=f"{mr_submitted} ({mr_submitted/total*100:.1f}%)" if total else "0")
        
        row += 1
        ws.cell(row=row, column=2, value="MRs Merged:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=f"{mr_merged} ({mr_merged/total*100:.1f}%)" if total else "0")
        
        # Interventions
        row += 2
        ws.cell(row=row, column=2, value="Interventions Needed:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=interventions_sent)
        
        # Add border around dashboard
        for r in range(2, row + 1):
            for c in [2, 3]:
                cell = ws.cell(row=r, column=c)
                cell.border = Styles.THIN_BORDER
    
    def _auto_fit_columns(self, ws) -> None:
        """Auto-fit column widths."""
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value or ""))
                    max_length = max(max_length, min(cell_length, 50))
                except:
                    pass
            
            ws.column_dimensions[column_letter].width = max_length + 2

