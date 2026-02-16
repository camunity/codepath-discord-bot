"""File processing services following SOLID principles.

This module provides extensible file processing capabilities:
- Abstract base class for processors (Open/Closed principle)
- CSV to Excel processor with styling
- File storage management
"""

import csv
import io
import os
from abc import ABC, abstractmethod
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, BinaryIO, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==================== Data Classes ====================

@dataclass
class ProcessingResult:
    """Result of a file processing operation."""
    success: bool
    output_data: Optional[bytes] = None
    output_filename: Optional[str] = None
    error_message: Optional[str] = None
    rows_processed: int = 0


@dataclass
class StoredFile:
    """Metadata for a stored file."""
    filename: str
    filepath: Path
    uploaded_at: datetime
    user_id: int
    file_type: str


# ==================== Abstract Base Class (Interface Segregation) ====================

class FileProcessor(ABC):
    """Abstract base class for file processors.
    
    Implements the Open/Closed principle - open for extension,
    closed for modification. New processors can be added by
    subclassing without modifying existing code.
    """
    
    @property
    @abstractmethod
    def input_type(self) -> str:
        """Return the expected input file type (e.g., 'csv', 'xlsx')."""
        pass
    
    @property
    @abstractmethod
    def output_type(self) -> str:
        """Return the output file type (e.g., 'xlsx', 'csv')."""
        pass
    
    @abstractmethod
    def process(self, data: bytes, options: Optional[Dict[str, Any]] = None) -> ProcessingResult:
        """Process the input data and return the result.
        
        Args:
            data: Raw bytes of the input file
            options: Optional processing options
            
        Returns:
            ProcessingResult with output data or error
        """
        pass


# ==================== Concrete Processors ====================

class CsvToExcelProcessor(FileProcessor):
    """Converts CSV files to styled Excel files.
    
    Single Responsibility: Only handles CSV to Excel conversion with styling.
    """
    
    # Default colors (can be overridden via options)
    DEFAULT_HEADER_COLOR = "4472C4"  # Blue header
    DEFAULT_ROW_COLOR = "92D050"     # Green rows
    DEFAULT_ALT_ROW_COLOR = "C6EFCE"  # Light green alternating
    
    @property
    def input_type(self) -> str:
        return "csv"
    
    @property
    def output_type(self) -> str:
        return "xlsx"
    
    def process(self, data: bytes, options: Optional[Dict[str, Any]] = None) -> ProcessingResult:
        """Convert CSV to styled Excel.
        
        Options:
            row_color: Hex color for data rows (default: green)
            header_color: Hex color for header row (default: blue)
            alternating: Use alternating row colors (default: False)
            alt_row_color: Hex color for alternating rows
        """
        options = options or {}
        
        try:
            # Decode CSV data
            text_data = data.decode('utf-8-sig')  # Handle BOM if present
            csv_reader = csv.reader(io.StringIO(text_data))
            rows = list(csv_reader)
            
            if not rows:
                return ProcessingResult(
                    success=False,
                    error_message="CSV file is empty"
                )
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Processed Data"
            
            # Define styles
            header_fill = PatternFill(
                start_color=options.get('header_color', self.DEFAULT_HEADER_COLOR),
                end_color=options.get('header_color', self.DEFAULT_HEADER_COLOR),
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            
            row_fill = PatternFill(
                start_color=options.get('row_color', self.DEFAULT_ROW_COLOR),
                end_color=options.get('row_color', self.DEFAULT_ROW_COLOR),
                fill_type="solid"
            )
            
            alt_row_fill = PatternFill(
                start_color=options.get('alt_row_color', self.DEFAULT_ALT_ROW_COLOR),
                end_color=options.get('alt_row_color', self.DEFAULT_ALT_ROW_COLOR),
                fill_type="solid"
            )
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            use_alternating = options.get('alternating', False)
            
            # Write data to worksheet
            for row_idx, row in enumerate(rows, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    if row_idx == 1:
                        # Header row
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        # Data rows
                        if use_alternating and row_idx % 2 == 0:
                            cell.fill = alt_row_fill
                        else:
                            cell.fill = row_fill
            
            # Auto-adjust column widths
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in column_cells:
                    try:
                        cell_length = len(str(cell.value or ""))
                        max_length = max(max_length, min(cell_length, 50))  # Cap at 50
                    except:
                        pass
                
                ws.column_dimensions[column_letter].width = max_length + 2
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return ProcessingResult(
                success=True,
                output_data=output.read(),
                output_filename="processed_data.xlsx",
                rows_processed=len(rows) - 1  # Exclude header
            )
            
        except UnicodeDecodeError as e:
            return ProcessingResult(
                success=False,
                error_message=f"Failed to decode CSV file: {e}"
            )
        except csv.Error as e:
            return ProcessingResult(
                success=False,
                error_message=f"Invalid CSV format: {e}"
            )
        except Exception as e:
            return ProcessingResult(
                success=False,
                error_message=f"Processing error: {e}"
            )


# ==================== File Storage Service (Single Responsibility) ====================

class FileStorageService:
    """Manages temporary file storage for processing.
    
    Single Responsibility: Only handles file storage and retrieval.
    Persists last file metadata to survive bot restarts.
    """
    
    def __init__(self, storage_dir: str = "data/uploads"):
        self.storage_dir = Path(storage_dir)
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        self._metadata_file = self.storage_dir / "_last_file.json"
        self._last_file: Optional[StoredFile] = self._load_last_file_metadata()
    
    def _load_last_file_metadata(self) -> Optional[StoredFile]:
        """Load last file metadata from disk."""
        if not self._metadata_file.exists():
            return None
        
        try:
            import json
            with open(self._metadata_file, 'r') as f:
                data = json.load(f)
            
            filepath = Path(data['filepath'])
            
            # Verify the file still exists
            if not filepath.exists():
                return None
            
            return StoredFile(
                filename=data['filename'],
                filepath=filepath,
                uploaded_at=datetime.fromisoformat(data['uploaded_at']),
                user_id=data['user_id'],
                file_type=data['file_type']
            )
        except Exception as e:
            print(f"[FileStorage] Failed to load metadata: {e}")
            return None
    
    def _save_last_file_metadata(self, stored: StoredFile) -> None:
        """Save last file metadata to disk."""
        try:
            import json
            data = {
                'filename': stored.filename,
                'filepath': str(stored.filepath),
                'uploaded_at': stored.uploaded_at.isoformat(),
                'user_id': stored.user_id,
                'file_type': stored.file_type
            }
            with open(self._metadata_file, 'w') as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            print(f"[FileStorage] Failed to save metadata: {e}")
    
    def store_file(self, filename: str, data: bytes, user_id: int) -> StoredFile:
        """Store a file and return its metadata."""
        # Generate unique filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_filename = "".join(c for c in filename if c.isalnum() or c in "._-")
        stored_name = f"{timestamp}_{safe_filename}"
        
        filepath = self.storage_dir / stored_name
        filepath.write_bytes(data)
        
        file_type = filename.rsplit('.', 1)[-1].lower() if '.' in filename else 'unknown'
        
        stored = StoredFile(
            filename=filename,
            filepath=filepath,
            uploaded_at=datetime.now(),
            user_id=user_id,
            file_type=file_type
        )
        
        self._last_file = stored
        self._save_last_file_metadata(stored)  # Persist to disk
        return stored
    
    def get_last_file(self) -> Optional[StoredFile]:
        """Get the most recently stored file."""
        # Verify the file still exists
        if self._last_file and not self._last_file.filepath.exists():
            self._last_file = None
        return self._last_file
    
    def read_file(self, stored_file: StoredFile) -> bytes:
        """Read and return file contents."""
        return stored_file.filepath.read_bytes()
    
    def cleanup_old_files(self, max_age_hours: int = 24) -> int:
        """Remove files older than max_age_hours. Returns count of deleted files."""
        deleted = 0
        cutoff = datetime.now().timestamp() - (max_age_hours * 3600)
        
        for filepath in self.storage_dir.iterdir():
            # Skip metadata file
            if filepath.name == "_last_file.json":
                continue
            if filepath.is_file() and filepath.stat().st_mtime < cutoff:
                filepath.unlink()
                deleted += 1
        
        return deleted


# ==================== Processor Registry (Dependency Inversion) ====================

class ProcessorRegistry:
    """Registry for file processors.
    
    Implements Dependency Inversion - high-level modules depend on
    this abstraction rather than concrete processors.
    """
    
    def __init__(self):
        self._processors: Dict[str, FileProcessor] = {}
    
    def register(self, name: str, processor: FileProcessor) -> None:
        """Register a processor by name."""
        self._processors[name] = processor
    
    def get(self, name: str) -> Optional[FileProcessor]:
        """Get a processor by name."""
        return self._processors.get(name)
    
    def get_by_input_type(self, input_type: str) -> Optional[FileProcessor]:
        """Get a processor that handles the given input type."""
        for processor in self._processors.values():
            if processor.input_type == input_type:
                return processor
        return None
    
    def list_processors(self) -> List[str]:
        """List all registered processor names."""
        return list(self._processors.keys())


# ==================== Default Registry Setup ====================

def create_default_registry() -> ProcessorRegistry:
    """Create a registry with default processors."""
    registry = ProcessorRegistry()
    registry.register('csv_to_excel', CsvToExcelProcessor())
    return registry

